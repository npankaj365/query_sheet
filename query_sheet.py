import cx_Oracle
from openpyxl import Workbook, load_workbook
connection = cx_Oracle.connect("PANKAJ", "PANKAJ", "192.168.50.103:1521/RMS14")
queries = ["""select 1 a, 2 b, 3 c
		from dual
		union
		select 4 a, 5 b, 6 c
		from dual"""
		,
		"""select 11 first, 12 second, 13 third from dual
		union
		select 21 first, 22 second, 23 third from dual"""]
cur = connection.cursor()
wb = Workbook()
ws = wb.active
r = 1
for query in queries:
	cur.execute(query)
	result = cur.fetchall()
	col_names = [row[0] for row in cur.description]

	print(col_names)
	c = 1
	for field_name in col_names:
		curr_cell = ws.cell(r, c)
		curr_cell.value = field_name
		c += 1
	r += 1
	for record in result:
		c = 1
		for item in record:
			curr_cell = ws.cell(r, c)
			#print(item, end="")
			curr_cell.value = item
			c += 1
		r += 1	
		c = 1
		print("\r")
wb.save('test.xlsx')
connection.close()


